import datetime
import glob
import re
import time
import threading
import os
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
import pandas as pd

begin_value = 'Sample'  # log数据开头第一个单词，一般为Sample

'''
目前测试可以使用的芯片列表如下：
BQ28Z610, BQ40Z50R2, SN27541
（同时也支持列表上没有芯片，只要log数据中模块名相同即可）
若没有相应的芯片类型，则可以把custom_type置为True，然后自定义custom_name的数据
'''
custom_type = False

'''
根据log数据中输出值的命名来修改custom_name中的值
如log上的时间名为ElapsedTime，则把custom_name中的TimeName改为ElapsedTime
'''
custom_name = ['TimeName', 'VoltageName', 'CurrentName', 'RSOCName', 'RCName', 'FCCName', 'TemperatureName']

# 时间显示线程使能
g_time_flag = 0

# 写入参数
g_author = ''
g_chr_voltage = 0
g_term_voltage = 0
g_fw_version = ''
g_project_name = ''


# 获取文件夹下所有log后缀的文件名
def get_file_name():
    filename = []
    for i in glob.glob(r'./*.log'):
        filename.append(i)
    for i in range(len(filename)):
        print(" %d : %s " % (i + 1, filename[i][2:]))
    file_num = input('\n输入文件编号：')
    while not file_num.isdigit() or (int(file_num) - 1) >= len(filename):
        file_num = input('输入错误，请重新输入：')
    return filename[int(file_num) - 1]


# 输出运行时间
def time_count():
    global g_time_flag
    while True:
        if g_time_flag == 0:
            time.sleep(0.1)
            continue
        elif g_time_flag == 1:
            start_time = time.time()
            while g_time_flag:
                print_time = time.time() - start_time
                print("\r%.2fs" % print_time, end='')
                time.sleep(0.01)


class BuildExcel:
    def __init__(self, log_name):
        self.file_path = log_name
        # 根据project name生成Excel文件名
        self.excel_path = ('./result/{0}-Battery-Cycle-Test-Curve-Rev{1}-{2:02}{3:02}{4}.xlsx'
                           .format(g_project_name, g_fw_version, datetime.datetime.now().month,
                                   datetime.datetime.now().day, datetime.datetime.now().year))
        self.log_name = None

        self.cycle_count = 0
        self.cycle_result = {}

        self.chr_current = 0
        self.disg_current = 0

    # 获取log数据中对应的模块名
    @staticmethod
    def get_module_name(line):
        if ('ElapsedTime' in line and 'Voltage' in line and 'Current' in line and 'RSOC' in line
                and 'RemCap' in line and 'FullChgCap' in line and 'Temperature' in line):

            return ['ElapsedTime', 'Voltage', 'Current', 'RSOC', 'RemCap', 'FullChgCap', 'Temperature']

        elif ('~Elapsed(s)' in line and 'Voltage' in line and 'AvgCurrent' in line and 'StateofChg' in line
              and 'RemCap' in line and 'FullChgCap' in line and 'Temperature' in line):

            return ['~Elapsed(s)', 'Voltage', 'AvgCurrent', 'StateofChg', 'RemCap', 'FullChgCap', 'Temperature']

        return False

    def log_to_excel(self):
        file = open(self.file_path, 'r')
        line = file.readlines()
        file.close()

        try:
            os.mkdir('./result/')
        except:
            pass

        # 当获取到begin_value后，定义这一行为开始行
        for i in range(len(line)):
            if begin_value in line[i]:
                begin_num = i
                break
            else:
                continue

        # 获取分隔符
        if ',' in line[begin_num]:
            delimiter = ','
        elif '\t' in line[begin_num]:
            delimiter = '\t'

        # 获取log数据中个模块的名称
        self.log_name = self.get_module_name(line[begin_num])

        # 若没有匹配数据，则检测有没有设置自定义模块
        if not self.log_name:
            if custom_type:
                self.log_name = custom_name
            else:
                return 'error1'

        i = begin_num
        # 根据分隔符将log数据分隔
        while i < len(line):
            line[i] = line[i].split(delimiter)
            i += 1

        # 生成新的二维数组
        new_line = line[begin_num:]

        len_data = len(new_line[1])

        for i in range(len(new_line)):
            # 定义首行数据，获取各数据模块的位置
            if i == 0:
                time_num = new_line[i].index(self.log_name[0])
                voltage_num = new_line[i].index(self.log_name[1])
                current_num = new_line[i].index(self.log_name[2])
                rsoc_num = new_line[i].index(self.log_name[3])
                rc_num = new_line[i].index(self.log_name[4])
                fcc_num = new_line[i].index(self.log_name[5])
                temp_num = new_line[i].index(self.log_name[6])
                new_line[i].extend([' ', 'Time', 'Voltage', 'Current', 'RSOC', 'RC', 'FCC', 'Temperature',
                                    ' ', 'Accumulated', 'Deviation', 'Fuel Gauge Deviation', 'Fuel Gauge Accuracy'])

            elif len(new_line[i]) >= len_data:
                ''' 将通讯错误引起的空白值改为上下数值的均值 '''
                if not new_line[i][time_num]:
                    if len(new_line[i+1]) < len_data or not new_line[i+1][time_num]:
                        temp_time = round((2*float(new_line[i-1][-7])-float(new_line[i-2][-7])), 6)
                    else:
                        temp_time = round((float(new_line[i-1][time_num])+float(new_line[i+1][time_num])/3600)/2, 6)
                else:
                    temp_time = round(float(new_line[i][time_num]) / 3600, 6)

                if not new_line[i][voltage_num]:
                    if len(new_line[i+1]) < len_data or not new_line[i+1][voltage_num]:
                        temp_vol = int(2*int(new_line[i-1][-6])-int(new_line[i-2][-6]))
                    else:
                        temp_vol = int((int(new_line[i - 1][-6]) + int(new_line[i + 1][voltage_num])) / 2)
                else:
                    temp_vol = int(new_line[i][voltage_num])

                if not new_line[i][current_num]:
                    if len(new_line[i+1]) < len_data or not new_line[i+1][current_num]:
                        temp_curr = abs(int(2*int(new_line[i-1][-5])-int(new_line[i-2][-5])))
                    else:
                        temp_curr = abs(int((int(new_line[i - 1][-5]) + abs(int(new_line[i + 1][current_num]))) / 2))
                else:
                    temp_curr = abs(int(new_line[i][current_num]))

                if not new_line[i][rsoc_num]:
                    if len(new_line[i+1]) < len_data or not new_line[i+1][rsoc_num]:
                        temp_rsoc = int(2*int(new_line[i-1][-4])-int(new_line[i-2][-4]))
                    else:
                        temp_rsoc = int((int(new_line[i-1][-4]) + int(new_line[i+1][rsoc_num]))/2)
                else:
                    temp_rsoc = int(new_line[i][rsoc_num])

                if not new_line[i][rc_num]:
                    if len(new_line[i+1]) < len_data or not new_line[i+1][rc_num]:
                        temp_rc = int(2*int(new_line[i-1][-3]) - int(new_line[i-2][-3]))
                    else:
                        temp_rc = int((int(new_line[i-1][-3]) + int(new_line[i+1][rc_num]))/2)
                else:
                    temp_rc = int(new_line[i][rc_num])

                if not new_line[i][fcc_num]:
                    if len(new_line[i+1]) < len_data or not new_line[i+1][fcc_num]:
                        temp_fcc = int(2*int(new_line[i-1][-2]) - int(new_line[i-2][-2]))
                    else:
                        temp_fcc = int((int(new_line[i-1][-2]) + int(new_line[i+1][fcc_num]))/2)
                else:
                    temp_fcc = int(new_line[i][fcc_num])

                if not new_line[i][temp_num]:
                    if len(new_line[i+1]) < len_data or not new_line[i+1][temp_num]:
                        temp_temp = float(2*float(new_line[i-1][-1]) - float(new_line[i-2][-1]))
                    else:
                        temp_temp = float((float(new_line[i-1][-1]) + float(new_line[i+1][temp_num]))/2)
                else:
                    temp_temp = float(new_line[i][temp_num])

                ''' 将添加特定数据 '''
                # 部分芯片通讯出现error时，新生成的数据的位置会被打乱，程序可以自动修复
                if re.search('error', new_line[i][-1], re.IGNORECASE) and '~Elapsed(s)' in new_line[0]:
                    pass
                else:
                    new_line[i].extend(' ')

                new_line[i].extend([temp_time, temp_vol, temp_curr, temp_rsoc, temp_rc, temp_fcc, temp_temp])

        # 计算容量
        cap_result = self.cap_accumulated(new_line)

        if not cap_result:
            return 'error2'

        # 将new_line的数据生成excel
        df = pd.DataFrame(new_line)
        df.to_excel(self.excel_path, header=None, index=False)

        return 'success'

    def cap_accumulated(self, line):
        # 获取新添加的数据的位置
        fcc_num = int(len(line[1])) - 2
        rc_num = fcc_num - 1
        rsoc_num = rc_num - 1
        current_num = rsoc_num - 1
        voltage_num = current_num - 1
        time_num = voltage_num - 1

        # 充放电标志位，chg_flag 和 disg_flag同时等于1时，自动计算容量程序才会开启
        # 即在检查到充电后，下一个放电操作才会计算容量
        chg_flag = 0
        disg_flag = 0

        i = 1
        while i < len(line) and len(line[i]) == len(line[1]):
            global g_term_voltage
            zero_num = 0
            term_num = 0

            if not -10 < line[i][current_num] < 10:
                begin_num = i

                while i < len(line) and len(line[i]) == len(line[1]):
                    if -10 < line[i][current_num] < 10:
                        end_num = i
                        break
                    else:
                        i += 1

                ''' 充放电判断 '''
                if line[begin_num][rsoc_num] < line[end_num][rsoc_num]:
                    # 充电
                    chg_flag = 1
                    chg_curr = line[round((end_num - begin_num) / 10) + begin_num][current_num]
                    self.chr_current = round(chg_curr / 100) / 10
                else:
                    # 放电
                    disg_flag = 1
                    disg_curr = line[round((end_num - begin_num) / 10) + begin_num][current_num]
                    self.disg_current = round(disg_curr / 100) / 10

                    # 判断为放电阶段时，记录下rc 0点和term点，用于计算容量差
                    for n in range(begin_num, end_num):
                        if line[n][rsoc_num] == 0 and zero_num == 0:
                            zero_num = n
                        if line[n][voltage_num] < g_term_voltage and term_num == 0:
                            term_num = n

                if chg_flag == 1 and disg_flag == 1:

                    self.cycle_count += 1

                    line[begin_num - 1].extend([' ', 0])
                    for n in range(begin_num, end_num):
                        # 容量计算公式
                        temp_cap = ((line[n][time_num] - line[n - 1][time_num]) *
                                    (line[n][current_num] + line[n - 1][current_num]) / 2 + line[n - 1][-1])

                        line[n].extend([' ', temp_cap])

                    ''' 一般情况下计算 '''
                    if zero_num != 0 and term_num != 0:
                        cap_dev = line[zero_num][-1] - line[term_num][-1]
                        cap_dev_percentage = cap_dev / line[term_num][-1]
                    else:
                        cap_dev = None
                        cap_dev_percentage = None

                    ''' 错误情况 '''
                    # 当没有term点时，返回错误
                    if term_num == 0:
                        return False

                    ''' 特殊情况1 '''
                    # 当RSOC瞬间跳为0时，检测上一时刻是否为1，若不是，cap_dev_percentage就为两个时刻的rsoc的差值
                    if cap_dev_percentage == 0:
                        if line[zero_num - 1][rsoc_num] > 1:
                            cap_dev_percentage = (line[zero_num][rsoc_num] - line[zero_num - 1][rsoc_num]) / 100

                    ''' 特殊情况2 '''
                    # 当rsoc没有出现0的情况下，cap_dev_percentage为term点的rsoc值
                    if zero_num == 0:
                        cap_dev = 0
                        cap_dev_percentage = line[term_num][rsoc_num] / 100

                    cap_percentage = line[term_num][-1] / line[begin_num][fcc_num]
                    if cap_percentage > 1:
                        cap_percentage = 1 / cap_percentage
                    line[term_num].extend(
                        [cap_dev, '{:.2%}'.format(cap_dev_percentage), '{:.2%}'.format(cap_percentage)])

                    if -0.06 <= cap_dev_percentage <= 0.06:
                        self.cycle_result['Cycle ' + str(self.cycle_count)] = ('{:.2%} PASS'.format(cap_dev_percentage))
                    else:
                        self.cycle_result['Cycle ' + str(self.cycle_count)] = ('{:.2%} FAIL'.format(cap_dev_percentage))

                    chg_flag = 0

            disg_flag = 0
            i += 1

        return True

    def print_chart(self):
        file = openpyxl.load_workbook(self.excel_path)
        sheet = file.active
        sheet.freeze_panes = 'A2'

        # 建立图表工作表
        chart_sheet = file.create_chartsheet('Chart1')

        chart1 = ScatterChart()
        chart2 = ScatterChart()

        # cycle test结果输出
        result_title = ''
        for i in self.cycle_result:
            result_title += '{0} : {1}   '.format(i, self.cycle_result[i])

        # 图表标题
        chart1.title = ('{0} Battery Pack Cycle-Test-Curve\n'
                        '\t\tF/W: {1},   Charge : {2}V/{3}A,   Discharge : {4}A\n'
                        '{5}\n'
                        '\t\t\t\t\tTested by:{6}'
                        .format(g_project_name, g_fw_version, g_chr_voltage,
                                self.chr_current, self.disg_current, result_title, g_author))

        # x轴为时间
        xvalue = Reference(sheet, min_row=2, min_col=sheet.max_column - 11,
                           max_row=sheet.max_row, max_col=sheet.max_column - 11)

        # 根据不同的数据建立Excel散点图系列
        for i in range(sheet.max_column - 10, sheet.max_column - 4):
            yvalue = Reference(sheet, min_row=1, min_col=i, max_row=sheet.max_row, max_col=i)

            series = Series(yvalue, xvalue, title_from_data=True)

            # RSOC和温度分配到另一个y轴
            if i == sheet.max_column - 8 or i == sheet.max_column - 5:
                chart2.append(value=series)
            else:
                chart1.append(value=series)

        # 关闭y轴框线
        chart1.x_axis.majorGridlines = None
        # 主y轴标题
        chart1.y_axis.title = 'Voltage(mV)/Current(mA)/RemCap(mAh)/FullChgCap(mAh)'

        # 副y轴标题
        chart2.y_axis.title = 'RSOC(%)/Temperature(\'C)'
        # 表示在x轴最大值处建立新的y轴
        chart2.y_axis.crosses = 'max'
        chart2.y_axis.axId = 200
        # 关闭y轴和x轴的框线
        chart2.y_axis.majorGridlines = None
        chart2.x_axis.majorGridlines = None

        chart1 += chart2
        chart_sheet.add_chart(chart1)

        file.save(self.excel_path)
        file.close()


def main():
    global g_time_flag
    global g_term_voltage
    global g_author
    global g_chr_voltage
    global g_fw_version
    global g_project_name

    print("####### 煲机数据自动处理工具V1.3 #######")
    file_name = get_file_name()
    g_project_name = input('请输入项目名称：')
    g_author = input('请输入作者：')
    g_fw_version = input('请输入软件版本：')
    g_chr_voltage = input('请输入充电电压 (V)：')
    g_term_voltage = int(input('输入term_voltage (mV)：'))

    build_excel = BuildExcel(file_name)
    time_count_thread = threading.Thread(target=time_count)
    time_count_thread.daemon = True
    time_count_thread.start()

    print('正在将log数据写入excel，请耐心等待...')
    g_time_flag = 1
    flag = build_excel.log_to_excel()
    g_time_flag = 0

    if flag == 'success':
        print('\n写入完成')
    elif flag == 'error1':
        print('\n暂时不支持该log格式')
        return False
    elif flag == 'error2':
        print('\n未搜索到term点电压，请检查参数是否输入正确')
        return False

    time.sleep(0.1)

    print('正在绘制图表，请耐心等待...')
    g_time_flag = 1
    build_excel.print_chart()
    g_time_flag = 0
    print('\n画图完成，文件保存在result文件夹下')


if __name__ == '__main__':
    main()
    input('按任意键退出')

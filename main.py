import glob
import re
import time
import threading
import os
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
import pandas as pd

begin_value = 'Sample'  # log数据开头第一个单词，一般为Sample

'''
目前测试可以使用的芯片列表如下：
BQ28Z610, BQ40Z50R2, SN27541
（同时也支持列表上没有芯片，只要log数据中模块名相同即可）
若没有相应的芯片类型，则可以把custom_type置为True，然后自定义custom_name的数据
'''
custom_type = False

'''
根据log数据中输出值的命名来修改type0中的值
如log上的时间名为ElapsedTime，则把custom_name中的TimeName改为ElapsedTime
'''
custom_name = ['TimeName', 'VoltageName', 'CurrentName', 'RSOCName', 'RCName', 'FCCName']

g_time_flag = 0
g_author = ''
g_chr_voltage = 0
g_term_voltage = 0
g_fw_version = ''


# 获取文件夹下所有log后缀的文件名
def get_file_name():
    filename = []

    for i in glob.glob(r'./*.log'):
        filename.append(i)

    for i in range(len(filename)):
        print(" %d : %s " % (i + 1, filename[i][2:]))

    file_num = input('\n输入文件编号：')

    while not file_num.isdigit() or (int(file_num) - 1) >= len(filename):
        file_num = input('输入错误，请重新输入：')

    return filename[int(file_num) - 1]


# 输出运行时间
def time_count():
    global g_time_flag
    while True:
        if g_time_flag == 0:
            time.sleep(0.1)
            continue
        elif g_time_flag == 1:
            start_time = time.time()
            while g_time_flag:
                print_time = time.time() - start_time
                print("\r%.2fs" % print_time, end='')
                time.sleep(0.01)


class BuildExcel:
    def __init__(self, log_name, ex_name):
        self.file_path = log_name
        self.excel_path = './result/' + ex_name + '.xlsx'
        self.log_name = None

        self.cycle_count = 0
        self.cycle_result = {}

        self.chr_current = 0
        self.disg_current = 0

    # 获取log数据中对应的模块名
    @staticmethod
    def get_module_name(line):
        if ('ElapsedTime' in line and 'Voltage' in line and 'Current' in line
                and 'RSOC' in line and 'RemCap' in line and 'FullChgCap' in line):

            return ['ElapsedTime', 'Voltage', 'Current', 'RSOC', 'RemCap', 'FullChgCap']

        elif ('~Elapsed(s)' in line and 'Voltage' in line and 'AvgCurrent' in line
              and 'StateofChg' in line and 'RemCap' in line and 'FullChgCap' in line):

            return ['~Elapsed(s)', 'Voltage', 'AvgCurrent', 'StateofChg', 'RemCap', 'FullChgCap']

        return False

    def log_to_excel(self):
        file = open(self.file_path, 'r')
        line = file.readlines()
        file.close()

        try:
            os.mkdir('./result/')
        except:
            pass

        for i in range(len(line)):
            if begin_value in line[i]:
                begin_num = i
                break
            else:
                continue

        if ',' in line[begin_num]:
            delimiter = ','
        elif '\t' in line[begin_num]:
            delimiter = '\t'

        i = begin_num

        self.log_name = self.get_module_name(line[begin_num])

        if not self.log_name:
            if custom_type:
                self.log_name = custom_name
            else:
                return False

        while i < len(line):
            line[i] = line[i].split(delimiter)
            i += 1

        new_line = line[begin_num:]

        for i in range(len(new_line)):
            if i == 0:
                time_num = new_line[i].index(self.log_name[0])
                voltage_num = new_line[i].index(self.log_name[1])
                current_num = new_line[i].index(self.log_name[2])
                rsoc_num = new_line[i].index(self.log_name[3])
                rc_num = new_line[i].index(self.log_name[4])
                fcc_num = new_line[i].index(self.log_name[5])
                new_line[i].extend([' ', 'Time', 'Voltage', 'Current', 'RSOC', 'RC', 'FCC',
                                    ' ', 'Accumulated', 'Deviation', 'Fuel Gauge Deviation', 'Fuel Gauge Accuracy'])

            else:
                ''' 将通讯错误引起的空白值改为0 '''
                if not new_line[i][time_num]:
                    new_line[i][time_num] = 0
                if not new_line[i][voltage_num]:
                    new_line[i][voltage_num] = 0
                if not new_line[i][current_num]:
                    new_line[i][current_num] = 0
                if not new_line[i][rsoc_num]:
                    new_line[i][rsoc_num] = 0
                if not new_line[i][rc_num]:
                    new_line[i][rc_num] = 0
                if not new_line[i][fcc_num]:
                    new_line[i][fcc_num] = 0

                if re.search('error', new_line[i][-1], re.IGNORECASE) and '~Elapsed(s)' in new_line[0]:
                    new_line[i].extend([round(float(new_line[i][time_num]) / 3600, 6),
                                        int(new_line[i][voltage_num]),
                                        abs(int(new_line[i][current_num])),
                                        int(new_line[i][rsoc_num]),
                                        int(new_line[i][rc_num]),
                                        int(new_line[i][fcc_num])])

                else:
                    new_line[i].extend([' ',
                                        round(float(new_line[i][time_num]) / 3600, 6),
                                        int(new_line[i][voltage_num]),
                                        abs(int(new_line[i][current_num])),
                                        int(new_line[i][rsoc_num]),
                                        int(new_line[i][rc_num]),
                                        int(new_line[i][fcc_num])])

        self.cap_accumulated(new_line)

        df = pd.DataFrame(new_line)
        df.to_excel(self.excel_path, header=None, index=False)

        return True

    def cap_accumulated(self, line):
        fcc_num = int(len(line[1])) - 1
        rc_num = fcc_num - 1
        rsoc_num = rc_num - 1
        current_num = rsoc_num - 1
        voltage_num = current_num - 1
        time_num = voltage_num - 1

        chg_flag = 0
        disg_flag = 0
        i = 1

        while i < len(line):
            global g_term_voltage
            zero_num = 0
            term_num = 0
            if not -10 < line[i][current_num] < 10:
                begin_num = i

                while True:
                    if -10 < line[i][current_num] < 10:
                        end_num = i
                        break
                    else:
                        i += 1

                ''' 充放电判断 '''
                if line[begin_num][rsoc_num] < line[end_num][rsoc_num]:
                    # 充电
                    chg_flag = 1
                    chg_curr = line[round((end_num - begin_num) / 10) + begin_num][current_num]
                    self.chr_current = round(chg_curr / 100) / 10
                else:
                    # 放电
                    disg_flag = 1
                    disg_curr = line[round((end_num - begin_num) / 10) + begin_num][current_num]
                    self.disg_current = round(disg_curr / 100) / 10

                    for n in range(begin_num, end_num):
                        if line[n][rsoc_num] == 0 and zero_num == 0:
                            zero_num = n
                        if line[n][voltage_num] < g_term_voltage and term_num == 0:
                            term_num = n

                if chg_flag == 1 and disg_flag == 1:

                    self.cycle_count += 1

                    line[begin_num - 1].extend([' ', 0])
                    for n in range(begin_num, end_num):
                        temp_cap = ((line[n][time_num] - line[n - 1][time_num]) *
                                    (line[n][current_num] + line[n - 1][current_num]) / 2 + line[n - 1][-1])

                        line[n].extend([' ', temp_cap])

                    cap_dev = line[zero_num][-1] - line[term_num][-1]
                    cap_dev_percentage = cap_dev / line[term_num][-1]
                    cap_percentage = line[term_num][-1] / line[begin_num][fcc_num]
                    if cap_percentage > 1:
                        cap_percentage = 1 / cap_percentage
                    line[term_num].extend(
                        [cap_dev, '{:.2%}'.format(cap_dev_percentage), '{:.2%}'.format(cap_percentage)])

                    if -0.06 <= cap_dev_percentage <= 0.06:
                        self.cycle_result['Cycle ' + str(self.cycle_count)] = ('{:.2%} PASS'.format(cap_dev_percentage))
                    else:
                        self.cycle_result['Cycle ' + str(self.cycle_count)] = ('{:.2%} FAIL'.format(cap_dev_percentage))

                    chg_flag = 0

            disg_flag = 0
            i += 1

    def print_chart(self):
        file = openpyxl.load_workbook(self.excel_path)
        sheet = file.active
        sheet.freeze_panes = 'A2'

        chart_sheet = file.create_chartsheet('Chart1')

        chart = ScatterChart()

        chart_rsoc = ScatterChart()

        result_title = ''
        for i in self.cycle_result:
            result_title += '{0} : {1}   '.format(i, self.cycle_result[i])

        chart.title = ('Project Name Cycle-Test-Curve\n'
                       '\t\tF/W: {0},   Charge : {1}V/{2}A,   Discharge : {3}A\n'
                       '{4}\n'
                       '\t\t\t\t\tTested by:{5}'.format(g_fw_version, g_chr_voltage, self.chr_current,
                                                        self.disg_current, result_title, g_author))

        xvalue = Reference(sheet, min_row=2, min_col=sheet.max_column - 10,
                           max_row=sheet.max_row, max_col=sheet.max_column - 10)

        for i in range(sheet.max_column - 9, sheet.max_column - 4):
            yvalue = Reference(sheet, min_row=1, min_col=i,
                               max_row=sheet.max_row, max_col=i)

            series = Series(yvalue, xvalue, title_from_data=True)
            if i == sheet.max_column - 7:
                chart_rsoc.append(value=series)
            else:
                chart.append(value=series)

        chart.x_axis.majorGridlines = None
        chart.y_axis.title = 'Voltage(mV)/Current(mA)/RemCap(mAh)/FullChgCap(mAh)'

        chart_rsoc.y_axis.title = 'RSOC(%)'
        chart_rsoc.y_axis.crosses = 'max'
        chart_rsoc.y_axis.axId = 200
        chart_rsoc.y_axis.majorGridlines = None
        chart_rsoc.x_axis.majorGridlines = None

        chart += chart_rsoc

        chart_sheet.add_chart(chart)

        file.save(self.excel_path)

        file.close()


def main():
    global g_time_flag
    global g_term_voltage
    global g_author
    global g_chr_voltage
    global g_fw_version

    file_name = get_file_name()
    excel_name = input('请输入导出Excel表格文件名（不需要添加后缀）：')
    g_author = input('请输入作者：')
    g_fw_version = input('请输入软件版本：')
    g_chr_voltage = input('请输入充电电压：')
    g_term_voltage = int(input('输入term_voltage (mV)：'))

    build_excel = BuildExcel(file_name, excel_name)
    time_count_thread = threading.Thread(target=time_count)
    time_count_thread.daemon = True
    time_count_thread.start()

    print('正在将log数据写入excel，请耐心等待...')
    g_time_flag = 1
    flag = build_excel.log_to_excel()
    g_time_flag = 0

    if flag:
        print('\n写入完成')
    else:
        print('\n不支持该log格式，请参考代码开头自定义数据名')
        return False

    time.sleep(0.1)

    print('正在绘制图表，请耐心等待...')
    g_time_flag = 1
    build_excel.print_chart()
    g_time_flag = 0
    print('\n画图完成，文件保存在result文件夹下')


if __name__ == '__main__':
    main()
    input('按任意键退出')
